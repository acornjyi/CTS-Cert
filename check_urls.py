"""
Dead URL Checker for DataON CoreTrustSeal Application_Update_250912.docx
Checks each URL with bot-bypass headers and writes results to Excel.
"""

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time
import re
from urllib.parse import urlparse

# ─── URL → Footnote mapping ───────────────────────────────────────────────────
# Each entry: (footnote_number, description, url)
FOOTNOTES = [
    (1,  "About DataON", "https://dataon.kisti.re.kr/intro/intro01.do?mm=MENU_00102&sm=MENU00144"),
    (2,  "About DataON (re3data)", "https://www.re3data.org/repository/r3d100013439"),
    (3,  "DataON Linkage and Services Agreement", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=31"),
    (4,  "Background and history of DataON development", "https://dataon.kisti.re.kr/intro/intro03.do?mm=MENU_00102&sm=MENU00146"),
    (5,  "DataON Services and Core Features", "https://dataon.kisti.re.kr/intro/intro02.do?mode=1&mm=MENU_00102&sm=MENU00145"),
    (6,  "DataON cooperated organizations", "https://dataon.kisti.re.kr/intro/intro07.do?mm=MENU_00102&sm=MENU00200"),
    (7,  "DataON Linkage and Services Agreement", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=31"),
    (8,  "Research Data Management Guidelines", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=26"),
    (9,  "Data Registration User Guide", "https://dataon.gitbook.io/dataon-user-guide/registration/how_to_register"),
    (10, "Data registration (personal information check)", "https://dataon.kisti.re.kr/dr/submit?mm=MENU_00088&sm=MENU00161"),
    (11, "Error Reporting (example)", "https://dataon.kisti.re.kr/search/view.do?mode=view&svcId=eb45ee61869dceceed2a08c4c97cd71e#errorModal"),
    (12, "Research Data Preservation Guidelines", "http://doi.org/10.22711/3"),
    (13, "Backup, Recovery, and Preservation Guidelines for DataON", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=25"),
    (14, "Identity Federation (KAFE)", "https://ds.kreonet.net/kafe?entityID=https%3A%2F%2Fdataon.kisti.re.kr%2Fsp%2Fspring&return=https%3A%2F%2Fdataon.kisti.re.kr%2Fsaml%2Flogin&returnIDParam=idp"),
    (15, "DOI API association assignment", "https://www.doi.or.kr/wordpress/"),
    (16, "NTIS Project Search API", "https://www.ntis.go.kr/rndopen/openApi/pjtSearch/"),
    (17, "ScienceON Article Search API", "https://scienceon.kisti.re.kr/main/mainForm.do"),
    (18, "KISTI Data Management Plan (DMP)", "https://dmp.kisti.re.kr/"),
    (19, "Integrated R&D Information System", "https://www.iris.go.kr/mbrs/srch/retrieveSrchNumInfo.do"),
    (20, "ORCID API", "https://info.orcid.org/documentation/features/public-api/"),
    (21, "Mission statement", "https://dataon.kisti.re.kr/intro/intro01.do?mm=MENU_00102&sm=MENU00144"),
    (22, "Research Data Management Regulations (draft)", "http://doi.org/10.22711/8"),
    (23, "Research Data Preservation Guidelines", "https://dataon.kisti.re.kr/data_mgnt_guideline_02.do"),
    (24, "Backup, Recovery, and Preservation Guidelines for DataON", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=25"),
    (25, "Data Registration Guidelines (disclosure and licensing)", "https://dataon.gitbook.io/dataon-user-guide/registration/how_to_register#7"),
    (26, "Research Data License Guidelines", "http://doi.org/10.22711/5"),
    (27, "Research Data Management Regulations (draft)", "http://doi.org/10.22711/8"),
    (28, "Terms of Use", "https://dataon.kisti.re.kr/layout/termsOfUse.do"),
    (29, "KISTI History", "https://www.kisti.re.kr/intro/pageView/15?t=1680151769423"),
    (30, "Act on Establishment of Government-funded Research Institutions", "https://www.law.go.kr/법령/과학기술분야정부출연연구기관등의설립ㆍ운영및육성에관한법률"),
    (31, "Research Data Preservation Guidelines", "https://dataon.kisti.re.kr/data_mgnt_guideline_02.do"),
    (32, "Research Data License Guidelines (Draft Standard Agreement)", "http://doi.org/10.22711/5"),
    (33, "DataON organization chart", "https://dataon.kisti.re.kr/intro/intro08.do"),
    (34, "Privacy Policy", "https://dataon.kisti.re.kr/layout/privacyPolicy.do"),
    (35, "Personal Information Protection Act", "https://www.law.go.kr/LSW/lsInfoP.do?chrClsCd=010203&lsiSeq=142563&viewCls=engLsInfoR&urlMode=engLsInfoR#0000"),
    (36, "Guidelines for Ensuring Research Ethics", "https://www.law.go.kr/행정규칙/연구윤리확보를위한지침/(263,20180717)"),
    (37, "Framework Act on Science and Technology", "https://www.law.go.kr/engLsSc.do?menuId=1&subMenuId=21&tabMenuId=117&query=%EA%B3%BC%ED%95%99%EA%B8%B0%EC%88%A0%EA%B8%B0%EB%B3%B8%EB%B2%95#"),
    (38, "Research Data License Guidelines", "http://doi.org/10.22711/5"),
    (39, "Research Data Ethics Guidelines", "http://doi.org/10.22711/6"),
    (40, "DataON Terms of Use", "http://doi.org/10.22711/1"),
    (41, "DataON Registration Guidelines", "https://dataon.gitbook.io/dataon-user-guide/registration/how_to_register#2."),
    (42, "Data Management Plan (DMP) Guidelines", "http://doi.org/10.22711/4"),
    (43, "Research Data Management Regulations (draft)", "http://doi.org/10.22711/8"),
    (44, "DataON organization chart", "https://dataon.kisti.re.kr/intro/intro08.do"),
    (45, "Research Data Management Guidelines", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=26"),
    (46, "DataON Service Operation and Management Guidelines", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=20"),
    (47, "KIRD", "https://www.kird.re.kr"),
    (48, "KISTI Science Data Education Center", "https://kacademy.kisti.re.kr/online-edu/free/620b7cf8be9d081049637166"),
    (49, "DataCite Member", "https://datacite.org/members.html"),
    (50, "IDW2022 session", "https://www.scidatacon.org/IDW-2022/sessions/293/"),
    (51, "RDA TAB", "https://www.rd-alliance.org/about-rda/our-leadership/rda-technical-advisory-board.html"),
    (52, "RDA OA-Member", "https://www.rd-alliance.org/oa-members"),
    (53, "KISTI organization chart", "https://www.kisti.re.kr/eng/about/pageView/248"),
    (54, "Operating Plan of Research Data Consultative Body", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=21"),
    (55, "Consultative Body Meeting Holding Details", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=32"),
    (56, "Institutional Research Data Repository Management Software Deployment", "http://doi.org/10.22711/12"),
    (60, "Research Data Registration Guidelines", "https://dataon.gitbook.io/dataon-user-guide/registration/how_to_register"),
    (61, "Research Data Preservation Guidelines", "http://doi.org/10.22711/3"),
    (62, "DataON and OAIS", "https://dataon.kisti.re.kr/intro/intro01.do?mm=MENU_00102&sm=MENU00144"),
    (63, "Research Data Management Guidelines", "http://doi.org/10.22711/2"),
    (64, "National Research and Development Information Processing Standards", "https://www.law.go.kr/%ED%96%89%EC%A0%95%EA%B7%9C%EC%B9%99/%EA%B5%AD%EA%B0%80%EC%97%B0%EA%B5%AC%EA%B0%9C%EB%B0%9C%EC%A0%95%EB%B3%B4%EC%B2%98%EB%A6%AC%EA%B8%B0%EC%A4%80/(2020-102,20201221)"),
    (65, "National Science and Technology Standard Classification System", "https://www.law.go.kr/%ED%96%89%EC%A0%95%EA%B7%9C%EC%B9%99/%EA%B5%AD%EA%B0%80%EA%B3%BC%ED%95%99%EA%B8%B0%EC%88%A0%ED%91%9C%EC%A4%80%EB%B6%84%EB%A5%98%EC%B2%B4%EA%B3%84"),
    (66, "Data Collection Scope", "https://dataon.kisti.re.kr/intro/intro01.do"),
    (67, "DataON's metadata schema", "https://dataon.kisti.re.kr/intro/intro09.do?mm=MENU_00102&sm=MENU00307"),
    (68, "DataON Registration Guidelines", "https://dataon.gitbook.io/dataon-user-guide/registration/how_to_register"),
    (69, "Catalog Schema Summary", "https://dataon.kisti.re.kr/catalog_summary.do"),
    (70, "Data Management Plan (DMP) Guidelines (1)", "http://doi.org/10.22711/4"),
    ("70b", "Data Management Plan (DMP) Guidelines (2)", "https://dmp.kisti.re.kr/main.do"),
    (71, "Research Data Management Guidelines", "http://doi.org/10.22711/2"),
    (72, "Research Data Preservation Guidelines", "http://doi.org/10.22711/3"),
    (73, "KISTI organization chart", "https://www.kisti.re.kr/eng/about/pageView/248"),
    (74, "Backup, Recovery, and Preservation Guidelines for DataON", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=25"),
    (75, "Research Data Management Guidelines", "https://dataon.kisti.re.kr/data_mgnt_guideline_01.do"),
    (76, "End user Terms and Conditions (DataON Terms of Use)", "http://doi.org/10.22711/1"),
    (77, "Research Data License Guidelines", "http://doi.org/10.22711/5"),
    (78, "Research Data Registration Guide (1)", "https://dataon.gitbook.io/dataon-user-guide/registration/how_to_register"),
    ("78b", "Research Data Ethics Guidelines", "http://doi.org/10.22711/6"),
    (79, "Research Data Registration Guide", "https://dataon.gitbook.io/dataon-user-guide/registration/how_to_register"),
    (80, "DataON Activity Guide", "https://dataon.gitbook.io/dataon-user-guide/mypage/manage"),
    (81, "Metadata error detection", "https://dataon.kisti.re.kr/DM/mgnt/main.do"),
    (82, "DataON Community Guide", "https://dataon.gitbook.io/dataon-user-guide/community"),
    (83, "Research Data Management Guideline_6 (Sharing)", "http://doi.org/10.22711/2"),
    (84, "Research Data Ethics Guidelines", "http://doi.org/10.22711/6"),
    (85, "Research Data Repository Utilization Guide", "http://doi.org/10.22711/9"),
    (86, "Research Data Management Guidelines_4 (Collecting)", "http://doi.org/10.22711/2"),
    (87, "Registration Guide", "https://dataon.gitbook.io/dataon-user-guide/registration/how_to_register"),
    (88, "CANVAS Guide", "https://dataon.gitbook.io/dataon-user-guide/canvas"),
    (89, "Data Management Plan (DMP) Guidelines (1)", "http://doi.org/10.22711/4"),
    ("89b", "Data Management Plan (DMP) Guidelines (2)", "https://dmp.kisti.re.kr/main.do"),
    (90, "Research Data Management Guidelines_5 (Publication)", "http://doi.org/10.22711/2"),
    (91, "Research Data Citation and Utilization Guidelines", "http://doi.org/10.22711/7"),
    (92, "User Guide (Research Data Search)", "https://dataon.gitbook.io/dataon-user-guide/search"),
    (93, "Metadata for Research Data Management Standard Reference", "https://www.tta.or.kr/tta/ttaSearchView.do?key=77&rep=1&searchStandardNo=TTAK.KO-10.0976&searchCate=TTAS"),
    (94, "DataON metadata XML schema", "https://dataon.kisti.re.kr/intro/intro09.do?mm=MENU_00102&sm=MENU00307"),
    (95, "Generate citation information", "https://dataon.kisti.re.kr/search/view.do?mode=view&svcId=4058c4325bd3a8c0d4e4490e16a0a0ce"),
    (96, "Backup, Recovery, and Preservation Guidelines for DataON", "https://dataon.kisti.re.kr/promVide/promVideView_R.do?pageIndex=1&promVideSn=25"),
]

# Domains belonging to DataON system
DATAON_DOMAINS = {
    "dataon.kisti.re.kr",
    "dataon.gitbook.io",
    "dmp.kisti.re.kr",
}

KISTI_DOMAINS = {
    "www.kisti.re.kr",
    "scienceon.kisti.re.kr",
    "kacademy.kisti.re.kr",
}

def is_dataon_page(url: str) -> str:
    try:
        host = urlparse(url).netloc.lower()
        if host in DATAON_DOMAINS:
            return "예 (DataON)"
        if host in KISTI_DOMAINS:
            return "예 (KISTI)"
        return "아니오"
    except Exception:
        return "확인불가"

def error_meaning(status_code, exception_msg=None):
    if exception_msg:
        if "ConnectionError" in exception_msg or "Failed to establish" in exception_msg:
            return "연결 실패 (서버 없음 또는 DNS 오류)"
        if "Timeout" in exception_msg or "timed out" in exception_msg.lower():
            return "연결 시간 초과"
        if "SSLError" in exception_msg:
            return "SSL 인증서 오류"
        if "TooManyRedirects" in exception_msg:
            return "리다이렉트 과다"
        return f"요청 오류: {exception_msg[:80]}"
    meanings = {
        200: "정상",
        201: "생성됨",
        204: "내용 없음 (정상)",
        301: "영구 이동 (정상 리다이렉트)",
        302: "임시 이동 (정상 리다이렉트)",
        400: "잘못된 요청",
        401: "인증 필요",
        403: "접근 금지 (봇차단 또는 권한 없음)",
        404: "페이지 없음 (Dead URL)",
        405: "허용되지 않는 메서드",
        410: "영구 삭제됨 (Dead URL)",
        429: "요청 과다 (봇차단)",
        500: "서버 내부 오류",
        502: "게이트웨이 오류",
        503: "서비스 이용불가",
        504: "게이트웨이 시간초과",
    }
    return meanings.get(status_code, f"HTTP {status_code}")

HEADERS_LIST = [
    # Primary: Chrome on Windows
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Cache-Control": "max-age=0",
    },
    # Fallback: Firefox on macOS
    {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:124.0) Gecko/20100101 Firefox/124.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
    },
]

def check_url(url: str, timeout: int = 15):
    """Return (status_code_or_None, bot_bypassed: bool, exception_msg_or_None, final_url)"""
    session = requests.Session()
    session.max_redirects = 10

    for i, headers in enumerate(HEADERS_LIST):
        try:
            resp = session.get(url, headers=headers, timeout=timeout, allow_redirects=True, verify=True)
            bypassed = (i > 0)  # second attempt succeeded
            return resp.status_code, bypassed, None, resp.url
        except requests.exceptions.SSLError:
            # retry without SSL verify
            try:
                resp = session.get(url, headers=headers, timeout=timeout, allow_redirects=True, verify=False)
                return resp.status_code, True, None, resp.url
            except Exception as e2:
                if i == len(HEADERS_LIST) - 1:
                    return None, False, type(e2).__name__ + ": " + str(e2), url
        except Exception as e:
            if i == len(HEADERS_LIST) - 1:
                return None, False, type(e).__name__ + ": " + str(e), url
            time.sleep(1)

    return None, False, "알 수 없는 오류", url


def result_label(status_code, exception_msg):
    if exception_msg:
        return "오류"
    if status_code is None:
        return "오류"
    if status_code in (200, 201, 204):
        return "정상"
    if status_code in (301, 302, 303, 307, 308):
        return "리다이렉트 (정상)"
    if status_code == 401:
        return "인증필요 (접근불가)"
    if status_code == 403:
        return "접근금지"
    if status_code in (404, 410):
        return "Dead URL"
    if status_code == 429:
        return "봇차단"
    if status_code >= 500:
        return "서버오류"
    return f"HTTP {status_code}"


def main():
    out_path = r"D:\업무\01 연구데이터센터\2026\01 CTS\DataON_URL_Check_Results.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "URL 검사 결과"

    headers = [
        "URL",
        "검사 결과",
        "서버 반환 코드",
        "엑셀 문서 내 URL 각주 번호",
        "봇차단 우회를 통한 정상 접근 여부",
        "오류의 의미",
        "DataON 내 페이지 여부",
    ]

    # Header style
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Row fills
    fill_ok     = PatternFill("solid", fgColor="C6EFCE")  # green
    fill_dead   = PatternFill("solid", fgColor="FFC7CE")  # red
    fill_warn   = PatternFill("solid", fgColor="FFEB9C")  # yellow
    fill_normal = PatternFill("solid", fgColor="FFFFFF")

    row_num = 2
    seen_urls = {}  # url → row first seen

    print(f"총 {len(FOOTNOTES)}개 항목 검사 시작...\n")

    for fn_num, description, url in FOOTNOTES:
        print(f"  [{fn_num}] {url[:70]}...", end=" ", flush=True)

        # Dedup: if same URL already checked, copy result
        if url in seen_urls:
            prev_row = seen_urls[url]
            prev_result   = ws.cell(row=prev_row, column=2).value
            prev_code     = ws.cell(row=prev_row, column=3).value
            prev_bypass   = ws.cell(row=prev_row, column=5).value
            prev_meaning  = ws.cell(row=prev_row, column=6).value
            prev_dataon   = ws.cell(row=prev_row, column=7).value
            row_data = [url, prev_result, prev_code, str(fn_num), prev_bypass, prev_meaning, prev_dataon]
            print(f"(캐시) → {prev_result}")
        else:
            status, bypassed, exc_msg, final_url = check_url(url)
            label  = result_label(status, exc_msg)
            meaning = error_meaning(status, exc_msg)
            bypass_label = "해당없음" if not exc_msg and status in (200, 201, 204) else ("예" if bypassed else "아니오")
            dataon_flag = is_dataon_page(url)

            code_str = str(status) if status is not None else "N/A"
            row_data = [url, label, code_str, str(fn_num), bypass_label, meaning, dataon_flag]
            seen_urls[url] = row_num
            print(f"→ {label} ({code_str})")
            time.sleep(0.8)  # polite delay

        ws.append(row_data)

        # Color the row based on result
        result = row_data[1]
        if "정상" in result or "리다이렉트" in result:
            row_fill = fill_ok
        elif "Dead" in result or "오류" in result:
            row_fill = fill_dead
        else:
            row_fill = fill_warn

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 1 or col == 6))

        row_num += 1

    # Column widths
    col_widths = [60, 20, 14, 22, 26, 34, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"\n완료! 결과 파일: {out_path}")


if __name__ == "__main__":
    main()
